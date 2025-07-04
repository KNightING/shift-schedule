import os
import google.generativeai as genai
import json
from datetime import date, timedelta, datetime
import random
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# --- V2 組態設定 ---
# 初始化 Gemini 模型
genai.configure(api_key=os.environ["GEMINI_API_KEY"])
model = genai.GenerativeModel("gemini-2.0-flash-lite")

# --- 職能與員工設定 ---
ROLES = ["前台", "外場", "吧台", "廚師", "管理"]
SHIFT_TYPES = ["早班", "午班", "晚班", "大夜班"]

# --- 特殊日期設定 ---
# 國定假日 (範例：2025年7月4日)
HOLIDAYS = [date(2025, 7, 4)]
# 不營業日 (範例：2025年7月7日)
NON_BUSINESS_DAYS = [date(2025, 7, 7)]



def generate_employees(num_employees=100):
    """產生具備隨機職能的員工資料"""
    employees = {}
    names = [f"員工{i}" for i in range(1, num_employees + 1)]
    for name in names:
        # 每個員工至少有一個職能，最多三個，模擬多能工
        num_roles = random.randint(1, 3)
        assigned_roles = random.sample(ROLES, num_roles)
        employees[name] = {"roles": assigned_roles}
    return employees


EMPLOYEES = generate_employees(100)
EMPLOYEE_NAMES = list(EMPLOYEES.keys())

# --- 固定班員工設定 ---
# 範例：設定員工1、員工2、員工3為固定班
FIXED_SHIFT_EMPLOYEES = {
    "員工1": {"shift": "早班", "role": "前台"},
    "員工2": {"shift": "午班", "role": "外場"},
    "員工3": {"shift": "晚班", "role": "廚師"},
}

# 更新 EMPLOYEE_NAMES 和 EMPLOYEES 字典，標記固定班員工
FIXED_SHIFT_EMPLOYEE_NAMES = list(FIXED_SHIFT_EMPLOYEES.keys())
for emp_name, details in FIXED_SHIFT_EMPLOYEES.items():
    if emp_name in EMPLOYEES:
        EMPLOYEES[emp_name]["is_fixed_shift"] = True
        EMPLOYEES[emp_name]["fixed_shift_type"] = details["shift"]
        # 確保固定班員工擁有其固定班的職能
        if details["role"] not in EMPLOYEES[emp_name]["roles"]:
            EMPLOYEES[emp_name]["roles"].append(details["role"])
    else:
        print(f"警告：固定班員工 {emp_name} 不存在於總員工列表中，已忽略設定。")

print(f"已設定 {len(FIXED_SHIFT_EMPLOYEES)} 位員工為固定班。")

# --- 班別與人力需求設定 ---
SHIFTS_PER_DAY_REQUIREMENTS = {
    "weekday": {
        "早班": {"前台": 2, "外場": 3, "吧台": 1, "廚師": 2, "管理": 1},
        "午班": {"前台": 3, "外場": 8, "吧台": 4, "廚師": 6, "管理": 1},
        "晚班": {"前台": 3, "外場": 8, "吧台": 4, "廚師": 6, "管理": 1},
        "大夜班": {"廚師": 1, "管理": 1},  # 夜間可能只有廚房和管理人員
    },
    "weekend": {
        "早班": {"前台": 3, "外場": 5, "吧台": 2, "廚師": 3, "管理": 1},
        "午班": {"前台": 3, "外場": 8, "吧台": 4, "廚師": 6, "管理": 1},
        "晚班": {"前台": 3, "外場": 8, "吧台": 4, "廚師": 6, "管理": 1},
        "大夜班": {"廚師": 2, "管理": 1},
    },
}

# --- 員工個人限制 ---
# 限制類型 1: 時間限制 (不能上某些班別)
constraint_employees = random.sample(EMPLOYEE_NAMES, 15)
EMPLOYEE_CONSTRAINTS = {
    # 只能上早班或午班
    constraint_employees[0]: ["早班", "午班"],
    constraint_employees[1]: ["早班", "午班"],
    # 只能上晚班或大夜班
    constraint_employees[4]: ["晚班", "大夜班"],
    # 不能上大夜班
    constraint_employees[6]: ["早班", "午班", "晚班"],
    constraint_employees[7]: ["早班", "午班", "晚班"],
}
print(f"已為 {len(EMPLOYEE_CONSTRAINTS)} 位員工設定特殊 [時間] 限制。")

# 將設定寫入檔案以供查閱
with open("employee_constraints.txt", "w", encoding="utf-8") as f:
    f.write("--- 當次執行所使用的員工職能設定 ---\n")
    f.write(json.dumps(EMPLOYEES, indent=2, ensure_ascii=False))
    f.write("\n--- 當次執行所使用的員工時間限制 ---\n")
    f.write(json.dumps(EMPLOYEE_CONSTRAINTS, indent=2, ensure_ascii=False))
print("詳細的員工職能與限制已寫入 employee_constraints.txt 檔案。")


# --- 共用函式 ---
def calculate_fairness_score(emp, stats, day_type, shift):
    """計算員工的公平性分數 (V2)"""
    score = stats[emp]["total_shifts"] * 0.1
    if day_type == "weekend":
        score += stats[emp]["weekend_shifts"] * 1.0
    if shift == "大夜班":
        score += stats[emp]["shift_counts"]["大夜班"] * 1.5
    # 新增：考慮各職能的分配公平性
    for role in ROLES:
        score += stats[emp]["role_counts"][role] * 0.2
    return score

def get_consecutive_work_days(employee, current_date, schedule):
    """計算員工在指定日期前連續上班的天數。"""
    consecutive_days = 0
    temp_date = current_date - timedelta(days=1)
    while True:
        date_str = temp_date.strftime("%Y-%m-%d")
        # 檢查日期是否在排班範圍內，並且該日期已經被處理過
        if date_str not in schedule:
            break

        worked_on_temp_date = False
        # 檢查員工是否在當天有任何班次
        for shift_data in schedule[date_str].values():
            for role_employees in shift_data.values():
                # 確保員工被排班且不是因為國定假日或不營業日而標記的「休」
                if employee in role_employees and "休" not in role_employees:
                    worked_on_temp_date = True
                    break
            if worked_on_temp_date:
                break

        if worked_on_temp_date:
            consecutive_days += 1
            temp_date -= timedelta(days=1)
        else:
            break  # 找到一天休息日，連續上班中斷

    return consecutive_days

def get_work_days_in_period(employee, current_date, schedule, period_days):
    """計算員工在指定日期前一個週期內 (period_days) 的上班天數。"""
    work_days = 0
    for i in range(period_days):
        check_date = current_date - timedelta(days=i)
        check_date_str = check_date.strftime("%Y-%m-%d")

        if check_date_str not in schedule:
            # 如果日期超出排班範圍，則停止檢查
            break

        worked_on_check_date = False
        for shift_data in schedule[check_date_str].values():
            for role_employees in shift_data.values():
                if employee in role_employees and "休" not in role_employees:
                    worked_on_check_date = True
                    break
            if worked_on_check_date:
                break

        if worked_on_check_date:
            work_days += 1
    return work_days


# --- 步驟 1: LLM 解析請求 (與 V1 相同) ---
def parse_requests_with_llm(raw_requests_text):
    """使用 Gemini 模型將非結構化的文字請求轉換為結構化的 JSON。"""
    prompt = f"""
    您是一位專業的 HR 排班助理。請仔細閱讀以下來自員工的排班請求文字，並將其轉換為一個結構化的 JSON 物件。
    JSON 物件必須包含一個 'leave_requests' 陣列。
    - 'leave_requests' 陣列中每個物件都應有 'employee' (員工姓名) 和 'date' (請假日期，格式為 YYYY-MM-DD)。
    請嚴格按照此格式輸出，不要有任何額外的說明文字。
    員工請求文字如下：
    ---
    {raw_requests_text}
    ---
    """
    print("--- 步驟 1: 正在呼叫 LLM 解析請假需求... ---")
    try:
        response = model.generate_content(prompt)
        clean_response = response.text.strip().replace("```json", "").replace("```", "")
        print(f"LLM 原始回應 (清理後): {clean_response}")
        return json.loads(clean_response)
    except (json.JSONDecodeError, AttributeError, Exception) as e:
        print(f"無法解析 LLM 的回應: {e}")
        print(
            f"LLM原始輸出: {response.text if 'response' in locals() else 'No response'}"
        )
        return {"leave_requests": []}


# --- 步驟 2: Python 演算法排班 (V2) ---
def create_schedule(start_date, end_date, leave_requests):
    """核心排班演算法 (V2: 支援職能)"""
    print("\n--- 步驟 2: 正在執行 Python 排班演算法 (V2)... ---")

    # V2 schedule structure: {date: {shift: {role: [employees]}}}
    schedule = {}
    # V2 stats structure
    stats = {
        emp: {
            "total_shifts": 0,
            "weekend_shifts": 0,
            "shift_counts": {st: 0 for st in SHIFT_TYPES},
            "role_counts": {r: 0 for r in ROLES},
        }
        for emp in EMPLOYEE_NAMES
    }

    leave_map = {}
    for req in leave_requests.get("leave_requests", []):
        if req["employee"] in EMPLOYEE_NAMES:
            leave_map.setdefault(req["date"], set()).add(req["employee"])

    delta = end_date - start_date
    for i in range(delta.days + 1):
        current_date = start_date + timedelta(days=i)
        date_str = current_date.strftime("%Y-%m-%d")
        schedule[date_str] = {st: {r: [] for r in ROLES} for st in SHIFT_TYPES}

        # 檢查是否為國定假日或不營業日
        if current_date in HOLIDAYS or current_date in NON_BUSINESS_DAYS:
            for shift in SHIFT_TYPES:
                for role in ROLES:
                    schedule[date_str][shift][role].append("休")  # 標記為休假
            # 將所有員工加入當天的請假列表，確保他們不會被排班
            for emp in EMPLOYEE_NAMES:
                leave_map.setdefault(date_str, set()).add(emp)
            continue  # 跳過當天的排班邏輯

        day_type = "weekend" if current_date.weekday() >= 5 else "weekday"
        requirements_today = SHIFTS_PER_DAY_REQUIREMENTS[day_type]
        on_leave_today = leave_map.get(date_str, set())

        # 一天內已排班的人員集合，避免重複排班
        assigned_today_flat = set()

        # --- 處理固定班員工 ---
        for emp_name in FIXED_SHIFT_EMPLOYEE_NAMES:
            if emp_name not in on_leave_today: # ��果員工沒有請假
                fixed_shift = EMPLOYEES[emp_name]["fixed_shift_type"]
                fixed_role = FIXED_SHIFT_EMPLOYEES[emp_name]["role"]
                
                # 確保固定班員工的職能符合其固定班別的需求
                if fixed_role in ROLES and fixed_shift in SHIFT_TYPES:
                    schedule[date_str][fixed_shift][fixed_role].append(emp_name)
                    assigned_today_flat.add(emp_name) # 將固定班員工加入已排班列表
                    
                    # 更新統計數據
                    stats[emp_name]["total_shifts"] += 1
                    stats[emp_name]["role_counts"][fixed_role] += 1
                    stats[emp_name]["shift_counts"][fixed_shift] += 1
                    if current_date.weekday() >= 5: # 判斷是否為週末
                        stats[emp_name]["weekend_shifts"] += 1
            else:
                # 如果固定班員工請假，則在排班表中標記為「休」
                for shift in SHIFT_TYPES:
                    for role in ROLES:
                        if emp_name in schedule[date_str][shift][role]:
                            schedule[date_str][shift][role].remove(emp_name)
                # 確保請假員工在當天不會被排班
                schedule[date_str][fixed_shift][fixed_role].append("休")

        # 取得前一天的排班資訊，用於「大夜班隔天不能上早班」規則
        previous_date = current_date - timedelta(days=1)
        previous_date_str = previous_date.strftime("%Y-%m-%d")
        employees_who_worked_night_shift_yesterday = set()
        if previous_date_str in schedule:
            night_shift_roles_prev_day = schedule[previous_date_str].get("大夜班", {})
            for role_employees in night_shift_roles_prev_day.values():
                employees_who_worked_night_shift_yesterday.update(role_employees)

        for shift, role_reqs in requirements_today.items():
            for role, count in role_reqs.items():
                for _ in range(count):
                    # 篩選可上班的員工
                    available_employees = [
                        emp
                        for emp in EMPLOYEE_NAMES
                        if emp not in on_leave_today
                        and emp not in assigned_today_flat
                        and emp not in FIXED_SHIFT_EMPLOYEE_NAMES # 排除固定班員工
                        and shift
                        in EMPLOYEE_CONSTRAINTS.get(emp, SHIFT_TYPES)  # 時間限制
                        and role in EMPLOYEES[emp]["roles"]  # 職能限制
                        # 新增規則: 大夜班隔天不能上早班
                        and not (shift == "早班" and emp in employees_who_worked_night_shift_yesterday)
                        # 新增規則: 不能連續上班6天 (即最多連續上班5天)
                        and get_consecutive_work_days(emp, current_date, schedule) < 5
                        # 新增規則: 雙周一定要休兩天假 (即14天內最多上班12天)
                        and get_work_days_in_period(emp, current_date, schedule, 14) < 12
                    ]

                    if not available_employees:
                        schedule[date_str][shift][role].append("!!人力不足!!")
                        continue

                    # 計算所有可用員工的公平性分數
                    employee_scores = [
                        (emp, calculate_fairness_score(emp, stats, day_type, shift))
                        for emp in available_employees
                    ]

                    # 找出最低分數
                    min_score = min(employee_scores, key=lambda x: x[1])[1]

                    # 收集所有達到最低分數的員工
                    best_employees_candidates = [
                        emp for emp, score in employee_scores if score == min_score
                    ]

                    # 從最佳候選人中隨機選擇一位
                    best_employee = random.choice(best_employees_candidates)

                    # 分派並更新數據
                    schedule[date_str][shift][role].append(best_employee)
                    assigned_today_flat.add(best_employee)
                    stats[best_employee]["total_shifts"] += 1
                    stats[best_employee]["role_counts"][role] += 1
                    stats[best_employee]["shift_counts"][shift] += 1
                    if day_type == "weekend":
                        stats[best_employee]["weekend_shifts"] += 1

    print("排班演算法完成！")
    return schedule, stats, leave_map


# --- 步驟 2.5: 最小幅度調整 (V2) ---
def find_and_assign_replacement(
    schedule, stats, leave_map, emp_on_leave, leave_date_str
):
    """為臨時請假的員工尋找最適合的代班人選 (V2)"""
    print(f"\n--- 正在為 {leave_date_str} 的 {emp_on_leave} 尋找代班人選... ---")

    # 1. 找出請假員工原本的班別和職能
    original_shift, original_role = None, None
    for shift, roles in schedule.get(leave_date_str, {}).items():
        for role, employees in roles.items():
            if emp_on_leave in employees:
                original_shift, original_role = shift, role
                break
        if original_shift:
            break

    if not original_shift or not original_role:
        print(f"錯誤：員工 {emp_on_leave} 在 {leave_date_str} 並沒有排班。")
        return schedule, stats, leave_map, False

    print(f"請假班別為: {original_shift} - {original_role}")

    # 2. 更新資料
    schedule[leave_date_str][original_shift][original_role].remove(emp_on_leave)
    leave_map.setdefault(leave_date_str, set()).add(emp_on_leave)

    # 3. 更新原員工統計數據
    day_is_weekend = datetime.strptime(leave_date_str, "%Y-%m-%d").weekday() >= 5
    stats[emp_on_leave]["total_shifts"] -= 1
    stats[emp_on_leave]["role_counts"][original_role] -= 1
    stats[emp_on_leave]["shift_counts"][original_shift] -= 1
    if day_is_weekend:
        stats[emp_on_leave]["weekend_shifts"] -= 1

    # 4. 尋找代班人選
    unavailable_today = set(leave_map.get(leave_date_str, set()))
    for shift_data in schedule.get(leave_date_str, {}).values():
        for role_employees in shift_data.values():
            unavailable_today.update(role_employees)

    potential_replacements = [
        emp
        for emp in EMPLOYEE_NAMES
        if emp not in unavailable_today
        and original_shift in EMPLOYEE_CONSTRAINTS.get(emp, SHIFT_TYPES)
        and original_role in EMPLOYEES[emp]["roles"]
    ]

    if not potential_replacements:
        print(
            f"警告：找不到符合資格的員工可以替補 {leave_date_str} 的 {original_shift} - {original_role}。"
        )
        schedule[leave_date_str][original_shift][original_role].append("!!人力不足!!")
        return schedule, stats, leave_map, False

    # 5. 找出最佳代班人
    day_type = "weekend" if day_is_weekend else "weekday"
    best_replacement = min(
        potential_replacements,
        key=lambda emp: calculate_fairness_score(emp, stats, day_type, original_shift),
    )

    # 6. 分派並更新數據
    schedule[leave_date_str][original_shift][original_role].append(best_replacement)
    stats[best_replacement]["total_shifts"] += 1
    stats[best_replacement]["role_counts"][original_role] += 1
    stats[best_replacement]["shift_counts"][original_shift] += 1
    if day_is_weekend:
        stats[best_replacement]["weekend_shifts"] += 1

    print(
        f"成功！由 {best_replacement} 替補 {emp_on_leave} 在 {leave_date_str} 的 {original_shift} - {original_role}。"
    )
    return schedule, stats, leave_map, True


# --- 步驟 3: LLM 生成報告 (V2) ---
def summarize_with_llm(schedule, stats):
    """將排班結果交給 LLM，生成易讀的報告 (V2)"""
    print("\n--- 步驟 3: 正在呼叫 LLM 生成摘要報告與通知 (V2)... ---")

    schedule_json = json.dumps(schedule, indent=2, ensure_ascii=False)
    stats_json = json.dumps(stats, indent=2, ensure_ascii=False)

    prompt = f"""
    您是一位經驗豐富的餐廳經理。這是一份剛剛產出的員工班表草案，以及相關的統計數據。班表現在區分了不同的「職能」(roles)。

    請根據這些資料，完成以下任務：

    1.  **生成摘要報告**:
        - 簡要說明班表的涵蓋期間。
        - 根據 'stats' 資料，進行公平性檢查。指出總班數、週末班或大夜班次數最多和最少的員工，並評估排班是否大致公平。
        - **關鍵任務**: 檢查 'schedule' 中是否有 '!!人力不足!!' 的標記。若有，請明確指出是哪個日期、哪個班別的哪個職能有人力缺口。
        - 根據 'stats' 中的 'role_counts'，分析是否有某些員工過度集中在特定職能，或者某些多技能員工沒有得到充分利用。

    2.  **草擬個人化通知 (範例)**:
        - 請為「員工1」和「員工2」這兩位員工，草擬一則簡短溫馨的排班通知訊息。訊息中需清楚列出他們各自的上班日期、班別、以及 **所負責的職能**。

    請用繁體中文、專業且友善的語氣來撰寫。

    班表資料 (JSON):
    {schedule_json}

    公平性統計資料 (JSON):
    {stats_json}
    """
    try:
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        return f"生成摘要時發生錯誤: {e}"


# --- 步驟 4: 匯出至 Excel (V2) ---
def export_to_excel(schedule, leave_map, filename="shift_schedule.xlsx"):
    """將排班結果匯出成一個格式化的 Excel 檔案 (V2.1: 總表 + 個人表)"""
    print(f"\n--- 步驟 4: 正在將班表匯出至 Excel ({filename})... ---")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 移除預設的工作表

    # --- 樣式設定 ---
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold_font = Font(bold=True)
    red_font = Font(color="FF0000", bold=True) # Added red_font definition
    leave_font = Font(color="FF0000", bold=True)
    leave_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    dates = sorted(schedule.keys())

    # --- 1. 建立總表 ---
    summary_sheet = wb.create_sheet(title="班表總表")

    # 寫入標頭
    summary_sheet.cell(row=1, column=1, value="員工").font = bold_font
    summary_sheet.cell(row=1, column=1).border = thin_border
    for col, date in enumerate(dates, start=2):
        cell = summary_sheet.cell(row=1, column=col, value=date)
        cell.font = bold_font
        cell.alignment = center_alignment
        # 判斷是否為六日，並設定紅色字體
        date_obj = datetime.strptime(date, "%Y-%m-%d").date()
        if date_obj.weekday() >= 5:
            cell.font = red_font
        cell.border = thin_border

    # 寫入員工列表及班表內容
    employee_row_map = {emp: i for i, emp in enumerate(EMPLOYEE_NAMES, start=2)}
    for emp, row_idx in employee_row_map.items():
        cell = summary_sheet.cell(row=row_idx, column=1, value=emp)
        cell.border = thin_border

    for date_idx, date in enumerate(dates, start=2):
        on_leave_today = leave_map.get(date, set())
        for emp, row_idx in employee_row_map.items():
            cell = summary_sheet.cell(row=row_idx, column=date_idx)
            cell.alignment = center_alignment
            cell.border = thin_border

            if emp in on_leave_today:
                cell.value = "休"
                cell.font = leave_font
                cell.fill = leave_fill
            else:
                assigned_info = []
                for shift, roles in schedule[date].items():
                    for role, employees in roles.items():
                        if emp in employees:
                            assigned_info.append(f"{shift}-{role}")
                if assigned_info:
                    cell.value = "\n".join(assigned_info)

    # 調整總表欄寬
    summary_sheet.column_dimensions["A"].width = 15
    for col_idx in range(2, len(dates) + 2):
        summary_sheet.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = 20

    # --- 2. 為每位員工建立個人班表 ---
    for emp in EMPLOYEE_NAMES:
        # 只有實際有班的員工才需要建立分頁
        if not any(
            emp in roles.get(r, [])
            for d in dates
            for s, roles in schedule.get(d, {}).items()
            for r in roles
        ):
            if not any(emp in leave_map.get(d, set()) for d in dates):
                continue

        emp_sheet = wb.create_sheet(title=f"{emp[:25]}")  # 避免工作表名稱過長

        # 寫入標頭
        emp_sheet.cell(row=1, column=1, value="班別").font = bold_font
        emp_sheet.cell(row=1, column=1).border = thin_border
        for col, date in enumerate(dates, start=2):
            cell = emp_sheet.cell(row=1, column=col, value=date)
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # 寫入班別列表
        shift_row_map = {shift: i for i, shift in enumerate(SHIFT_TYPES, start=2)}
        for shift, row_idx in shift_row_map.items():
            cell = emp_sheet.cell(row=row_idx, column=1, value=shift)
            cell.border = thin_border

        # 填入個人班表資料
        for date_idx, date in enumerate(dates, start=2):
            on_leave_today = leave_map.get(date, set())
            if emp in on_leave_today:
                for row_idx in shift_row_map.values():
                    cell = emp_sheet.cell(row=row_idx, column=date_idx)
                    cell.value = "休"
                    cell.font = leave_font
                    cell.alignment = center_alignment
                    cell.fill = leave_fill
                    cell.border = thin_border
            else:
                for shift, row_idx in shift_row_map.items():
                    assigned_role = None
                    for role, employees in schedule[date].get(shift, {}).items():
                        if emp in employees:
                            assigned_role = role
                            break

                    cell = emp_sheet.cell(row=row_idx, column=date_idx)
                    if assigned_role:
                        cell.value = assigned_role
                    cell.alignment = center_alignment
                    cell.border = thin_border

        # 調整個人表欄寬
        emp_sheet.column_dimensions["A"].width = 12
        for col_idx in range(2, len(dates) + 2):
            emp_sheet.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = 15

    wb.save(filename)
    print("Excel 檔案匯出成功！")


# --- 主程式執行流程 ---
if __name__ == "__main__":
    raw_requests = """
    To HR:
    我是員工1，我因為家裡有事，想要在 2025年7月15日 請假一天，謝謝。
    員工2: 你好，我想請 2025-07-18，再麻煩了。
    員工3: 老闆，7/15 我沒辦法上班喔。
    """

    # 步驟 1
    parsed_data = parse_requests_with_llm(raw_requests)

    # 步驟 2
    start_date = date(2025, 7, 1)
    end_date = date(2025, 7, 31)
    final_schedule, final_stats, leave_map = create_schedule(
        start_date, end_date, parsed_data
    )

    print("\n--- 初版班表已生成 ---")
    initial_filename = "shift_schedule_initial.xlsx"
    export_to_excel(final_schedule, leave_map, filename=initial_filename)
    print(f"已產生初版班表 Excel 檔案: {initial_filename}")

    # 互動式調整迴圈 (暫時停用以進行非互動式測試)
    # while True:
    #     adjust = input("\n是否需要進行手動調班？ (請輸入 'y' 進行調整，或直接按 Enter 結束): ").lower()
    #     if adjust != "y":
    #         break

    #     emp_to_replace = input(f"請輸入要請假的員工姓名 (例如 員工1): ")
    #     if emp_to_replace not in EMPLOYEE_NAMES:
    #         print("錯誤：找不到該員工。")
    #         continue

    #     date_to_replace = input("請輸入請假日期 (格式 YYYY-MM-DD): ")
    #     try:
    #         datetime.strptime(date_to_replace, "%Y-%m-%d")
    #     except ValueError:
    #         print("錯誤：日期格式不正確。")
    #         continue

    #     final_schedule, final_stats, leave_map, success = find_and_assign_replacement(
    #         final_schedule, final_stats, leave_map, emp_to_replace, date_to_replace
    #     )
    #     if success:
    #         print(f"班表已更新。")

    print("\n--- 所有調整已完成 ---")

    # 步驟 3 (暫時停用 目前不需要摘要報告)
    # summary_report = summarize_with_llm(final_schedule, final_stats)
    # print("\n\n" + "=" * 50)
    # print("     由 Gemini Pro 生成的最終排班摘要報告與通知")
    # print("=" * 50)
    # print(summary_report)

    # 步驟 4 (最終版)
    final_filename = "shift_schedule_final.xlsx"
    export_to_excel(final_schedule, leave_map, filename=final_filename)
    print(f"已產生最終版班表 Excel 檔案: {final_filename}")
